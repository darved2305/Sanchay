"""Pure-function tests for USP heuristic fallbacks (no LLM, no live DB)."""

from __future__ import annotations

import asyncio
import io
import uuid

from pydantic import SecretStr

from app.connectors.google import decrypt_token, encrypt_token, fixture_harvest, sign_oauth_state, verify_oauth_state
from app.core.config import Settings
from app.services.admin_request import build_multi_faculty_output, detect_header_row, resolve_row_for_faculty, summarize_gaps
from app.services.any_form import compute_coverage, detect_fields_xlsx, fill_workbook, resolve_field, resolve_field_with_llm
from app.services.evidence_match import find_evidence_matches
from app.services.career import evaluate_career_rules, match_opportunities
from app.services.cv_import import _heuristic_extract, extract_cv_activities
from app.services.llm import LLMProvider
from app.services.lor import draft_letter, polish_letter
from app.services.network import rank_candidates, score_candidate
from app.services.quick_add import _heuristic_parse, parse_quick_add
from app.services.reconstruct import is_academic_signal, run_pipeline
from app.services.teaching_change import (
    SnapshotFile,
    deterministic_change_descriptions,
    diff_snapshots,
    has_meaningful_changes,
    summarize_changes,
)

from .support import UNCONFIGURED_LLM


def test_quick_add_heuristic_detects_category_and_duration() -> None:
    parsed = _heuristic_parse("Conducted a 2-hour seminar on GenAI today for TE IT")
    assert parsed["category"] == "seminar"
    assert parsed["duration_hours"] == 2.0


def test_quick_add_heuristic_falls_back_to_other() -> None:
    parsed = _heuristic_parse("Went for a walk")
    assert parsed["category"] == "other"


def test_quick_add_without_llm_key_uses_heuristic() -> None:
    settings = Settings(**UNCONFIGURED_LLM)
    result = asyncio.run(parse_quick_add("Attended a workshop on cloud computing", LLMProvider(settings)))
    assert result["category"] == "workshop_fdp"
    assert result["academic_year"]
    assert result["start_date"]


def test_cv_heuristic_extracts_keyworded_lines() -> None:
    cv_text = (
        "Professor of Computer Science\n"
        "Attended FDP on Generative AI, IIT Bombay, 2024\n"
        "Published a paper in IEEE Transactions, 2023\n"
        "Just a random unrelated line about hobbies\n"
    )
    drafts = _heuristic_extract(cv_text)
    categories = {d["category"] for d in drafts}
    assert "workshop_fdp" in categories
    assert "publication" in categories
    assert len(drafts) == 2


def test_cv_import_without_llm_key_uses_heuristic() -> None:
    settings = Settings(**UNCONFIGURED_LLM)
    cv_text = "Delivered an invited talk at NIT Surat, 2022\n"
    drafts = asyncio.run(extract_cv_activities(cv_text, LLMProvider(settings)))
    assert len(drafts) == 1
    assert drafts[0]["category"] == "invited_talk"
    assert drafts[0]["academic_year"] == "2022-23"


CAREER_RULES = [
    {"key": "research", "label": "Research publications", "categories": ["publication"], "min_count": 2},
    {"key": "service", "label": "Service", "categories": ["committee"], "min_count": 3},
]


def _activity(activity_id: str, category: str, evidence_status: str = "attached") -> dict:
    return {"id": activity_id, "category": category, "evidence_status": evidence_status}


def test_career_rules_count_and_flag_unmet_thresholds() -> None:
    activities = [
        _activity("1", "publication"),
        _activity("2", "publication"),
        _activity("3", "committee"),
    ]
    result = evaluate_career_rules(CAREER_RULES, activities)
    research_rule = next(r for r in result["rules"] if r["key"] == "research")
    service_rule = next(r for r in result["rules"] if r["key"] == "service")
    assert research_rule["count"] == 2 and research_rule["satisfied"] is True
    assert service_rule["count"] == 1 and service_rule["satisfied"] is False and service_rule["gap"] == 2
    assert result["satisfied"] is False


def test_career_rules_evidence_completeness_counts_each_activity_once() -> None:
    activities = [_activity("1", "publication", "attached"), _activity("2", "publication", "pending")]
    result = evaluate_career_rules(CAREER_RULES, activities)
    assert result["evidence_completeness"] == 50.0


def test_career_rules_no_rules_means_not_satisfied() -> None:
    assert evaluate_career_rules([], [])["satisfied"] is False


def test_opportunity_matching_only_targets_unmet_rules_with_a_reason() -> None:
    rules = evaluate_career_rules(CAREER_RULES, [_activity("1", "publication"), _activity("2", "publication")])["rules"]
    opportunities = [
        {"id": "opp-service", "kind": "committee", "tags": ["service"], "deadline": "2026-09-01"},
        {"id": "opp-research", "kind": "publication", "tags": ["research"], "deadline": None},
    ]
    recs = match_opportunities(rules, opportunities)
    assert len(recs) == 1
    assert recs[0]["opportunity_id"] == "opp-service"
    assert "Service" in recs[0]["reason"]
    assert "2026-09-01" in recs[0]["reason"]


def test_reconstruct_fixture_filters_out_non_academic_signals() -> None:
    harvested = fixture_harvest()
    academic = [item for item in harvested if is_academic_signal(item)]
    non_academic_titles = {item.title for item in harvested} - {item.title for item in academic}
    assert "Weekend trip photos" in non_academic_titles


def test_reconstruct_fixture_correlates_the_signature_ieee_talk_across_three_sources() -> None:
    candidates = asyncio.run(run_pipeline(fixture_harvest()))
    ieee_candidate = next((c for c in candidates if "IEEE" in c["title"]), None)
    assert ieee_candidate is not None
    assert ieee_candidate["category"] == "invited_talk"
    source_types = {source["source_type"] for source in ieee_candidate["sources"]}
    assert source_types == {"gmail", "google_calendar", "google_drive"}
    assert ieee_candidate["confidence"] == 0.9


def test_reconstruct_fixture_keeps_independent_signals_as_separate_candidates() -> None:
    candidates = asyncio.run(run_pipeline(fixture_harvest()))
    all_snippets = " ".join(source["snippet"].lower() for candidate in candidates for source in candidate["sources"])
    assert "manuscript" in all_snippets
    # The reviewer thank-you and the FDP day-3 event must not merge into the IEEE cluster.
    assert len(candidates) >= 3


def _make_xlsx(rows: list[list[str | None]]) -> bytes:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_any_form_detects_label_value_pairs() -> None:
    content = _make_xlsx([["Faculty Name:", ""], ["Department:", ""], ["Random note", "not a field"]])
    fields = detect_fields_xlsx(content)
    labels = {f.label for f in fields}
    assert "Faculty Name" in labels
    assert "Department" in labels
    assert "Random note" not in labels


def test_any_form_resolves_known_labels_from_context() -> None:
    context = {"full_name": "Dr. Ananya Sharma", "category_counts": {"publication": 5}}
    value, confidence, resolve_status = resolve_field("Faculty Name", context)
    assert value == "Dr. Ananya Sharma" and resolve_status == "auto_filled" and confidence > 0.5
    value, confidence, resolve_status = resolve_field("Number of Publications", context)
    assert value == 5 and resolve_status == "auto_filled"


def test_any_form_unknown_label_needs_new_info() -> None:
    value, confidence, resolve_status = resolve_field("Favorite programming language", {})
    assert value is None and resolve_status == "needs_new_info"


def test_any_form_llm_fallback_without_key_still_needs_new_info() -> None:
    settings = Settings(**UNCONFIGURED_LLM)
    value, confidence, resolve_status = asyncio.run(
        resolve_field_with_llm("Favorite programming language", {}, LLMProvider(settings))
    )
    assert value is None and resolve_status == "needs_new_info"


def test_any_form_llm_fallback_skips_llm_when_already_resolved() -> None:
    settings = Settings(**UNCONFIGURED_LLM)
    context = {"full_name": "Dr. Ananya Sharma"}
    value, confidence, resolve_status = asyncio.run(resolve_field_with_llm("Faculty Name", context, LLMProvider(settings)))
    assert value == "Dr. Ananya Sharma" and resolve_status == "auto_filled"


def test_any_form_coverage_counts_match_field_statuses() -> None:
    fields = [
        {"status": "auto_filled"}, {"status": "auto_filled"}, {"status": "needs_confirmation"}, {"status": "needs_new_info"},
    ]
    coverage = compute_coverage(fields)
    assert coverage == {"fields_detected": 4, "fields_auto_filled": 2, "fields_needs_confirmation": 1, "fields_needs_new_info": 1}


def test_any_form_fill_workbook_preserves_other_cells() -> None:
    content = _make_xlsx([["Faculty Name:", ""], ["Untouched", "Value"]])
    filled = fill_workbook(content, [{"sheet": "Sheet", "value_cell": "B1", "value": "Dr. Ananya Sharma"}])
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(filled))
    sheet = workbook.active
    assert sheet["B1"].value == "Dr. Ananya Sharma"
    assert sheet["A2"].value == "Untouched"
    assert sheet["B2"].value == "Value"


def test_evidence_match_suggests_the_right_pending_activity() -> None:
    evidence = {"file_name": "IIT_Bombay_FDP_certificate.pdf", "extracted_title": "Certificate of Participation", "organization": "IIT Bombay"}
    pending = [
        {"id": "a1", "title": "FDP on Generative AI", "organization": "IIT Bombay"},
        {"id": "a2", "title": "Committee meeting minutes", "organization": "Internal"},
    ]
    matches = find_evidence_matches(evidence, pending)
    assert len(matches) == 1
    assert matches[0]["activity"]["id"] == "a1"


def test_evidence_match_returns_nothing_below_threshold() -> None:
    evidence = {"file_name": "random_scan.pdf"}
    pending = [{"id": "a1", "title": "Completely unrelated seminar", "organization": "Somewhere"}]
    assert find_evidence_matches(evidence, pending) == []


def test_admin_request_detects_table_header_not_label_colon_shape() -> None:
    content = _make_xlsx([
        ["Department CSE - FDP Participation 2023-26"],
        ["Faculty Name", "Employee Code", "Number of FDPs"],
        ["", "", ""],
    ])
    result = detect_header_row(content)
    assert result is not None
    row_index, labels, columns = result
    assert row_index == 2
    assert labels == ["Faculty Name", "Employee Code", "Number of FDPs"]
    assert columns == [1, 2, 3]


def test_admin_request_header_detection_preserves_blank_column_positions() -> None:
    content = _make_xlsx([["Name", None, "Publications"], ["", "", ""]])
    result = detect_header_row(content)
    assert result is not None
    row_index, labels, columns = result
    assert row_index == 1
    assert labels == ["Name", "Publications"]
    assert columns == [1, 3]


def test_admin_request_resolves_one_row_per_faculty_from_their_own_context() -> None:
    labels = ["Faculty Name", "Number of Publications", "Favorite color"]
    context_a = {"full_name": "Dr. A", "category_counts": {"publication": 3}}
    context_b = {"full_name": "Dr. B", "category_counts": {"publication": 0}}
    row_a = resolve_row_for_faculty(labels, context_a)
    row_b = resolve_row_for_faculty(labels, context_b)
    assert row_a["Faculty Name"]["value"] == "Dr. A" and row_a["Number of Publications"]["value"] == 3
    assert row_b["Faculty Name"]["value"] == "Dr. B" and row_b["Number of Publications"]["value"] == 0
    assert row_a["Favorite color"]["status"] == "needs_new_info"


def test_admin_request_summarizes_faculty_with_gaps() -> None:
    rows = [
        {"fields": {"a": {"status": "auto_filled"}, "b": {"status": "auto_filled"}}},
        {"fields": {"a": {"status": "auto_filled"}, "b": {"status": "needs_new_info"}}},
    ]
    summary = summarize_gaps(rows)
    assert summary == {"faculty_count": 2, "faculty_with_gaps": 1}


def test_admin_request_writes_one_row_per_faculty_below_header() -> None:
    content = _make_xlsx([["Faculty Name", "Publications"], ["", ""]])
    labels = ["Faculty Name", "Publications"]
    faculty_rows = [
        {"fields": {"Faculty Name": {"value": "Dr. A"}, "Publications": {"value": 3}}},
        {"fields": {"Faculty Name": {"value": "Dr. B"}, "Publications": {"value": 5}}},
    ]
    output = build_multi_faculty_output(content, 1, labels, faculty_rows)
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(output))
    sheet = workbook.active
    assert sheet["A1"].value == "Faculty Name" and sheet["B1"].value == "Publications"
    assert sheet["A2"].value == "Dr. A" and sheet["B2"].value == 3
    assert sheet["A3"].value == "Dr. B" and sheet["B3"].value == 5


def test_admin_request_writes_values_into_real_columns_when_header_has_blanks() -> None:
    content = _make_xlsx([["Name", None, "Publications"], ["", "", ""]])
    detected = detect_header_row(content)
    assert detected is not None
    _, labels, columns = detected
    faculty_rows = [
        {"fields": {"Name": {"value": "Dr. A"}, "Publications": {"value": 7}}},
    ]
    output = build_multi_faculty_output(content, 1, labels, faculty_rows, columns)
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(output))
    sheet = workbook.active
    assert sheet["A1"].value == "Name" and sheet["C1"].value == "Publications"
    assert sheet["A2"].value == "Dr. A"
    assert sheet["B2"].value in (None, "")
    assert sheet["C2"].value == 7


def test_teaching_change_detects_added_removed_and_modified_files() -> None:
    files_a = [
        SnapshotFile("syllabus.pdf", "hash1", "Week 1: intro\nWeek 2: basics"),
        SnapshotFile("lab1.pdf", "hashlab1", "Lab exercise A"),
        SnapshotFile("old_notes.pdf", "hashold", "deprecated"),
    ]
    files_b = [
        SnapshotFile("syllabus.pdf", "hash2", "Week 1: intro\nWeek 2: basics\nWeek 3: advanced topics"),
        SnapshotFile("lab1.pdf", "hashlab1", "Lab exercise A"),
        SnapshotFile("lab2_new.pdf", "hashlab2", "New lab on cloud deployment"),
    ]
    diff = diff_snapshots(files_a, files_b)
    assert diff["added"] == ["lab2_new.pdf"]
    assert diff["removed"] == ["old_notes.pdf"]
    assert diff["unchanged"] == ["lab1.pdf"]
    assert len(diff["changed"]) == 1
    assert diff["changed"][0]["file_name"] == "syllabus.pdf"
    assert diff["changed"][0]["lines_added"] == 1
    assert has_meaningful_changes(diff) is True


def test_teaching_change_no_diff_is_honest_not_changed() -> None:
    files = [SnapshotFile("syllabus.pdf", "samehash", "same content")]
    diff = diff_snapshots(files, files)
    assert diff["added"] == [] and diff["removed"] == [] and diff["changed"] == []
    assert has_meaningful_changes(diff) is False


def test_teaching_change_deterministic_descriptions_reflect_the_diff() -> None:
    diff = {"added": ["lab2.pdf"], "removed": ["old.pdf"], "changed": [{"file_name": "syllabus.pdf", "lines_added": 5, "lines_removed": 1}], "unchanged": []}
    descriptions = deterministic_change_descriptions(diff)
    types = {d["change_type"] for d in descriptions}
    assert types == {"material_added", "material_removed", "material_modified"}
    assert any("lab2.pdf" in d["description"] for d in descriptions)


def test_teaching_change_summarize_without_llm_key_falls_back_to_deterministic() -> None:
    settings = Settings(**UNCONFIGURED_LLM)
    diff = {"added": ["lab2.pdf"], "removed": [], "changed": [], "unchanged": []}
    result = asyncio.run(summarize_changes(diff, LLMProvider(settings)))
    assert len(result) == 1
    assert result[0]["change_type"] == "material_added"


def test_teaching_change_empty_diff_summarizes_to_nothing() -> None:
    settings = Settings(**UNCONFIGURED_LLM)
    diff = {"added": [], "removed": [], "changed": [], "unchanged": ["syllabus.pdf"]}
    result = asyncio.run(summarize_changes(diff, LLMProvider(settings)))
    assert result == []


LOR_FACTS = {
    "student_name": "Rohan Mehta",
    "purpose": "ms",
    "faculty_name": "Dr. Ananya Sharma",
    "designation": "Associate Professor",
    "institution_name": "Vidyanagar Institute of Technology",
    "links": [{"relationship": "project guide", "course_or_project": "Capstone Project", "start_date": "August 2024", "end_date": "May 2025", "notes": "He consistently demonstrated initiative."}],
    "achievements": [{"title": "Best Capstone Award", "description": "Awarded for the top final-year project", "achieved_on": "May 2025"}],
}


def test_lor_draft_includes_only_the_given_facts() -> None:
    letter = draft_letter(LOR_FACTS)
    assert "Rohan Mehta" in letter
    assert "project guide" in letter
    assert "Capstone Project" in letter
    assert "Best Capstone Award" in letter
    assert "Dr. Ananya Sharma" in letter
    assert "admission to a Master's program" in letter


def test_lor_draft_never_invents_achievements_when_none_recorded() -> None:
    facts = {**LOR_FACTS, "achievements": []}
    letter = draft_letter(facts)
    assert "demonstrated strong ability" not in letter
    assert "Rohan Mehta" in letter


def test_lor_polish_without_llm_key_returns_deterministic_draft_unchanged() -> None:
    settings = Settings(**UNCONFIGURED_LLM)
    draft = draft_letter(LOR_FACTS)
    result = asyncio.run(polish_letter(draft, LLMProvider(settings)))
    assert result == draft


OAUTH_SETTINGS = Settings(google_oauth_client_secret=SecretStr("test-client-secret-value"))
OTHER_OAUTH_SETTINGS = Settings(google_oauth_client_secret=SecretStr("a-completely-different-secret"))


def test_oauth_state_round_trips_user_id_and_provider() -> None:
    user_id = uuid.uuid4()
    state = sign_oauth_state(OAUTH_SETTINGS, user_id, "gmail")
    result = verify_oauth_state(OAUTH_SETTINGS, state)
    assert result == (user_id, "gmail")


def test_oauth_state_rejects_tampering() -> None:
    user_id = uuid.uuid4()
    state = sign_oauth_state(OAUTH_SETTINGS, user_id, "gmail")
    assert verify_oauth_state(OTHER_OAUTH_SETTINGS, state) is None


def test_oauth_state_rejects_garbage_input() -> None:
    assert verify_oauth_state(OAUTH_SETTINGS, "not-a-real-state-token") is None


def test_oauth_token_encryption_round_trips() -> None:
    ciphertext = encrypt_token(OAUTH_SETTINGS, "ya29.real-looking-access-token")
    assert ciphertext != "ya29.real-looking-access-token"
    assert decrypt_token(OAUTH_SETTINGS, ciphertext) == "ya29.real-looking-access-token"


def test_oauth_token_decryption_fails_with_wrong_key() -> None:
    ciphertext = encrypt_token(OAUTH_SETTINGS, "secret-token")
    assert decrypt_token(OTHER_OAUTH_SETTINGS, ciphertext) is None


SEEKER = {"research_interests": ["medical imaging", "deep learning"], "expertise": ["python"], "department_name": "Computer Science"}


def test_network_score_reflects_shared_research_interests() -> None:
    candidate = {"id": "c1", "research_interests": ["Medical Imaging", "Robotics"], "expertise": [], "department_name": "Electronics"}
    result = score_candidate(SEEKER, candidate)
    assert result["score"] > 0
    assert any("medical imaging" in r for r in result["reasons"])


def test_network_score_explains_phd_supervisor_intent() -> None:
    candidate = {"id": "c1", "research_interests": [], "expertise": [], "department_name": None, "accepting_phd_inquiries": True}
    result = score_candidate(SEEKER, candidate, intent="phd_supervisor")
    assert any("PhD" in r for r in result["reasons"])


def test_network_score_zero_overlap_has_no_reasons() -> None:
    candidate = {"id": "c1", "research_interests": [], "expertise": [], "department_name": "Chemistry"}
    result = score_candidate(SEEKER, candidate)
    assert result["score"] == 0 and result["reasons"] == []


def test_network_rank_drops_unexplainable_candidates() -> None:
    candidates = [
        {"id": "match", "research_interests": ["medical imaging"], "expertise": [], "department_name": None},
        {"id": "no-match", "research_interests": ["astrophysics"], "expertise": [], "department_name": "Physics"},
    ]
    ranked = rank_candidates(SEEKER, candidates)
    assert len(ranked) == 1
    assert ranked[0]["profile_id"] == "match"


def test_network_rank_orders_by_score_descending() -> None:
    candidates = [
        {"id": "one-match", "research_interests": ["deep learning"], "expertise": [], "department_name": None},
        {"id": "two-match", "research_interests": ["deep learning", "medical imaging"], "expertise": ["python"], "department_name": "Computer Science"},
    ]
    ranked = rank_candidates(SEEKER, candidates)
    assert [r["profile_id"] for r in ranked] == ["two-match", "one-match"]
