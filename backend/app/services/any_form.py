"""USP 2 — Any Form Assistant: XLSX field detection, mapping, and fill.

Deterministic-first: field *detection* (which cells are label:value pairs)
and field *resolution* (which canonical value fills a recognized label) are
both plain rules over the workbook structure and a keyword table -- no LLM
call is required for the common case. An LLM, when configured, is only used
by the caller as a fallback mapper for labels this keyword table doesn't
recognize (see api/forms.py), and its result still goes through the same
confidence/needs-confirmation bucketing.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from .llm import LLMProvider


@dataclass
class DetectedField:
    sheet: str
    label: str
    label_cell: str
    value_cell: str


def detect_fields_xlsx(content: bytes) -> list[DetectedField]:
    """Find label:value regions: a text cell whose adjacent cell (right, then
    below) is empty is treated as a field waiting to be filled. This covers
    the common "Name: ____" / two-column form layout without guessing at
    arbitrary table structures.
    """

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    fields: list[DetectedField] = []
    for sheet in workbook.worksheets:
        max_row = min(sheet.max_row or 0, 500)
        max_col = min(sheet.max_column or 0, 50)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                text = str(cell.value).strip() if cell.value is not None else ""
                if not text or not text.endswith(":"):
                    continue
                label = text.rstrip(":").strip()
                if not label or len(label) > 150:
                    continue
                right_cell = sheet.cell(row=row, column=col + 1) if col < max_col else None
                below_cell = sheet.cell(row=row + 1, column=col) if row < max_row else None
                target = None
                if right_cell is not None and (right_cell.value is None or str(right_cell.value).strip() == ""):
                    target = right_cell
                elif below_cell is not None and (below_cell.value is None or str(below_cell.value).strip() == ""):
                    target = below_cell
                if target is not None:
                    fields.append(DetectedField(
                        sheet=sheet.title,
                        label=label,
                        label_cell=f"{get_column_letter(col)}{row}",
                        value_cell=f"{get_column_letter(target.column)}{target.row}",
                    ))
    return fields


CANONICAL_RESOLVERS: list[tuple[str, str]] = [
    ("employee code", "employee_code"),
    ("emp code", "employee_code"),
    ("full name", "full_name"),
    ("faculty name", "full_name"),
    ("name", "full_name"),
    ("email", "email"),
    ("phone", "phone"),
    ("mobile", "phone"),
    ("designation", "designation"),
    ("department", "department_name"),
    ("institution", "institution_name"),
    ("college", "institution_name"),
    ("orcid", "orcid_id"),
    ("date of joining", "date_joined"),
    ("joining date", "date_joined"),
    ("academic year", "current_academic_year"),
    ("qualification", "qualifications_summary"),
    ("number of publications", "count:publication"),
    ("no of publications", "count:publication"),
    ("publications", "count:publication"),
    ("number of fdps", "count:workshop_fdp"),
    ("fdps attended", "count:workshop_fdp"),
    ("workshops attended", "count:workshop_fdp"),
    ("committees", "count:committee"),
    ("invited talks", "count:invited_talk"),
    ("patents", "count:patent"),
    ("awards", "count:award"),
]


def _match_canonical_path(label: str) -> str | None:
    lowered = label.lower()
    for keyword, canonical_path in CANONICAL_RESOLVERS:
        if keyword in lowered:
            return canonical_path
    return None


def resolve_field(label: str, context: dict[str, Any]) -> tuple[Any, float, str]:
    """Return (value, confidence, status) for one detected label.

    status is one of: auto_filled, needs_confirmation, needs_new_info.
    """

    canonical_path = _match_canonical_path(label)
    if canonical_path is None:
        return None, 0.0, "needs_new_info"
    if canonical_path.startswith("count:"):
        category = canonical_path.split(":", 1)[1]
        value = context.get("category_counts", {}).get(category, 0)
        return value, 0.95, "auto_filled"
    value = context.get(canonical_path)
    if value in (None, ""):
        return None, 0.3, "needs_confirmation"
    return value, 0.9, "auto_filled"


_CANONICAL_PATHS = sorted({path for _, path in CANONICAL_RESOLVERS})

_FIELD_MAPPING_SCHEMA = {
    "type": "object",
    "properties": {"canonical_path": {"type": "string", "enum": [*_CANONICAL_PATHS, "none"]}},
    "required": ["canonical_path"],
}


def _resolve_canonical_path(canonical_path: str, context: dict[str, Any], confidence: float) -> tuple[Any, float, str]:
    if canonical_path.startswith("count:"):
        category = canonical_path.split(":", 1)[1]
        return context.get("category_counts", {}).get(category, 0), confidence, "auto_filled"
    value = context.get(canonical_path)
    if value in (None, ""):
        return None, 0.3, "needs_confirmation"
    return value, confidence, "auto_filled"


async def resolve_field_with_llm(label: str, context: dict[str, Any], llm: LLMProvider) -> tuple[Any, float, str]:
    """Same contract as ``resolve_field``, but falls back to an LLM mapper when
    the keyword table doesn't recognize the label -- e.g. "Institute Email ID"
    instead of "Email". The LLM only ever picks which of the known canonical
    fields (the exact same set ``resolve_field`` uses) a label refers to; it
    never invents a value itself, so a wrong guess degrades to
    'needs_new_info' at worst, never a fabricated fact.
    """

    value, confidence, resolve_status = resolve_field(label, context)
    if resolve_status != "needs_new_info" or not llm.configured:
        return value, confidence, resolve_status
    result = await llm.extract_structured(
        instruction=(
            "This is a label from an institutional form. Map it to the single closest matching "
            "canonical field from the enum, or 'none' if nothing genuinely matches. Do not guess "
            "loosely -- 'none' is the correct answer for anything ambiguous or unrelated."
        ),
        source_text=label,
        json_schema=_FIELD_MAPPING_SCHEMA,
        schema_name="form_field_mapping",
    )
    canonical_path = result.get("canonical_path") if result else None
    if not canonical_path or canonical_path == "none" or canonical_path not in _CANONICAL_PATHS:
        return value, confidence, resolve_status
    return _resolve_canonical_path(canonical_path, context, confidence=0.7)


def compute_coverage(fields: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "fields_detected": len(fields),
        "fields_auto_filled": sum(1 for f in fields if f["status"] == "auto_filled"),
        "fields_needs_confirmation": sum(1 for f in fields if f["status"] == "needs_confirmation"),
        "fields_needs_new_info": sum(1 for f in fields if f["status"] == "needs_new_info"),
    }


def fill_workbook(content: bytes, filled_fields: list[dict[str, Any]]) -> bytes:
    """Write resolved values back into their original cells, preserving
    every other cell, style, formula, and sheet untouched."""

    workbook = openpyxl.load_workbook(io.BytesIO(content))
    by_sheet: dict[str, list[dict[str, Any]]] = {}
    for field in filled_fields:
        if field.get("value") in (None, ""):
            continue
        by_sheet.setdefault(field["sheet"], []).append(field)
    for sheet_name, sheet_fields in by_sheet.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for field in sheet_fields:
            sheet[field["value_cell"]] = field["value"]
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
