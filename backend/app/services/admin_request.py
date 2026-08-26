"""USP 6 — Admin Request Autopilot: the multi-faculty sibling of Any Form.

Where Any Form (services/any_form.py) fills one entity's form, this fills a
table with one row per faculty member. It reuses the exact same
label -> canonical-field resolver (``resolve_field``) so "Number of
Publications" means the same thing whether one professor is filling their
own form or an admin is filling a department-wide roster.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl

from .any_form import resolve_field

MAX_HEADER_SCAN_ROWS = 10


def detect_header_row(content: bytes) -> tuple[int, list[str], list[int]] | None:
    """Find the first row that looks like a table header: at least two
    non-empty text cells, none of which end in ':' (that's Any Form's
    label:value shape, not a table header). Returns the 1-based row index,
    the label texts, and each label's real 1-based column index (blank
    spacer columns are skipped here but must not shift later writes)."""

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.worksheets[0]
    for row_index in range(1, min(sheet.max_row or 0, MAX_HEADER_SCAN_ROWS) + 1):
        labels: list[str] = []
        columns: list[int] = []
        for col_index in range(1, (sheet.max_column or 0) + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            text = str(value).strip() if value is not None else ""
            if text:
                labels.append(text)
                columns.append(col_index)
        non_colon_labels = [label for label in labels if not label.endswith(":")]
        if len(non_colon_labels) >= 2:
            return row_index, labels, columns
    return None


def resolve_row_for_faculty(labels: list[str], context: dict[str, Any]) -> dict[str, Any]:
    """Resolve one output row: {label: {value, status}} for a single faculty member."""

    row: dict[str, Any] = {}
    for label in labels:
        value, _confidence, status = resolve_field(label, context)
        row[label] = {"value": value, "status": status}
    return row


def build_multi_faculty_output(content: bytes, header_row_index: int, labels: list[str], faculty_rows: list[dict[str, Any]], columns: list[int] | None = None) -> bytes:
    """Write one row per faculty below the header, preserving the header and
    everything above it exactly as uploaded. ``columns`` holds each label's
    real 1-based sheet column; when omitted the labels are assumed to be
    contiguous starting at column A."""

    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.worksheets[0]
    write_row = header_row_index + 1
    label_columns = columns if columns is not None else list(range(1, len(labels) + 1))
    for faculty_row in faculty_rows:
        for col_index, label in zip(label_columns, labels):
            resolved = faculty_row["fields"].get(label, {})
            value = resolved.get("value")
            sheet.cell(row=write_row, column=col_index, value=value if value not in (None, "") else "")
        write_row += 1
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def summarize_gaps(faculty_rows: list[dict[str, Any]]) -> dict[str, Any]:
    faculty_with_gaps = sum(
        1 for row in faculty_rows
        if any(field["status"] != "auto_filled" for field in row["fields"].values())
    )
    return {"faculty_count": len(faculty_rows), "faculty_with_gaps": faculty_with_gaps}
